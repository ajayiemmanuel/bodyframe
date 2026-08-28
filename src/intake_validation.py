"""Reproducible BODYFRAME Stage 1 file-intake validation.

The module reads the master registry, inspects registered files, applies the
established Stage 1 rules, and returns dictionaries matching FileManifestTable.
It does not modify canonical workbooks, manifests, or image files.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import mimetypes
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from PIL import Image


INSPECTION_METHOD = "python_pillow_v1"

MANIFEST_FIELDS = [
    "asset_id",
    "source_image_id",
    "parent_asset_id",
    "is_original",
    "variant_type",
    "variant_level",
    "filename",
    "relative_filepath",
    "registry_match_check",
    "file_exists",
    "file_readable",
    "file_extension",
    "detected_format",
    "extension_format_match",
    "mime_type",
    "file_size_bytes",
    "width_px",
    "height_px",
    "aspect_ratio",
    "orientation",
    "bit_depth",
    "color_mode",
    "icc_profile",
    "exif_available",
    "date_captured",
    "camera_make",
    "camera_model",
    "lens_model",
    "focal_length_mm",
    "iso",
    "aperture_f_number",
    "shutter_speed_seconds",
    "sha256",
    "exact_duplicate",
    "duplicate_group_id",
    "filename_rule_pass",
    "folder_rule_pass",
    "required_metadata_complete_pct",
    "intake_status",
    "failure_reason",
    "inspection_timestamp",
    "inspection_method",
    "notes",
]

# These nine fields reproduce the established 4/9 and 6/9 completeness values
# for the controlled missing and unreadable assets. EXIF is intentionally absent.
REQUIRED_METADATA_FIELDS = [
    "asset_id",
    "source_image_id",
    "filename",
    "relative_filepath",
    "file_size_bytes",
    "width_px",
    "height_px",
    "detected_format",
    "sha256",
]

CONTROLLED_FAILURE_IDS = [
    "BF_S001_0001_BADNAME",
    "BF_S001_0001_DUPLICATE",
    "BF_S001_0002_WRONGFOLDER",
    "BF_S001_0003_MISSING",
    "BF_S001_0003_UNREADABLE",
]

CLEAN_PILOT_IDS = [
    "BF_S001_0001_ORIGINAL",
    "BF_S001_0002_ORIGINAL",
    "BF_S001_0003_ORIGINAL",
    "BF_S001_0001_JPEG_Q40",
]

ORIGINAL_FILENAME_RE = re.compile(r"^BF_S\d{3}_\d{4}\.jpg$")
DERIVATIVE_FILENAME_RE = re.compile(
    r"^BF_S\d{3}_\d{4}__[a-z0-9]+(?:_[a-z0-9]+)*\.jpg$"
)
EXPANSION_ASSET_RE = re.compile(r"^BF_S(?:00[2-9]|010|011)_\d{4}_ORIGINAL$")


def _read_sheet_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read a rectangular worksheet into dictionaries without modifying it."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows)]
        records = []
        for values in rows:
            if not any(value is not None and value != "" for value in values):
                continue
            records.append(dict(zip(headers, values, strict=False)))
        return records
    finally:
        workbook.close()


def load_registry_assets(registry_path: Path) -> list[dict[str, Any]]:
    """Load all nonblank asset rows from the registry's assets sheet."""
    return [
        row
        for row in _read_sheet_records(registry_path, "assets")
        if row.get("asset_id")
    ]


def load_canonical_manifest(manifest_path: Path) -> list[dict[str, Any]]:
    """Load canonical FileManifestTable values for reproducibility comparison."""
    return [
        row
        for row in _read_sheet_records(manifest_path, "file_manifest")
        if row.get("asset_id")
    ]


def resolve_registered_path(project_root: Path, relative_filepath: Any) -> Path:
    """Resolve a registry path and reject paths that escape the project root."""
    relative = PurePosixPath(str(relative_filepath or ""))
    resolved_root = project_root.resolve()
    resolved = (resolved_root / Path(*relative.parts)).resolve()
    if resolved != resolved_root and resolved_root not in resolved.parents:
        raise ValueError(f"Registered path escapes project root: {relative_filepath}")
    return resolved


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    """Calculate a file SHA-256 without loading the entire file into memory."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _exif_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    try:
        return float(value)
    except (TypeError, ValueError, ZeroDivisionError):
        return str(value)


def _bit_depth(mode: str) -> int:
    if mode == "1":
        return 1
    if "16" in mode:
        return 16
    return 8


def _aspect_ratio(width: int, height: int) -> str:
    divisor = math.gcd(width, height)
    return f"{width // divisor}:{height // divisor}"


def filename_rule_passes(asset: Mapping[str, Any]) -> bool:
    """Apply the existing BODYFRAME original/derivative filename patterns."""
    filename = str(asset.get("filename") or "")
    source_image_id = str(asset.get("source_image_id") or "")
    if str(asset.get("is_original") or "").lower() == "yes":
        return bool(ORIGINAL_FILENAME_RE.fullmatch(filename)) and filename == (
            source_image_id + ".jpg"
        )
    return bool(DERIVATIVE_FILENAME_RE.fullmatch(filename)) and filename.startswith(
        source_image_id + "__"
    )


def folder_rule_passes(asset: Mapping[str, Any]) -> bool:
    """Require originals and derivatives to use their established folders."""
    filename = str(asset.get("filename") or "")
    relative = str(asset.get("relative_filepath") or "").replace("\\", "/")
    if str(asset.get("is_original") or "").lower() == "yes":
        return relative == f"data/private_images/originals/{filename}"
    return relative == f"data/processed_variants/{filename}"


def inspect_asset(
    asset: Mapping[str, Any],
    project_root: Path,
    registry_id_counts: Mapping[str, int],
    inspection_timestamp: str | None = None,
) -> dict[str, Any]:
    """Inspect one registered asset without making any filesystem changes."""
    asset_id = str(asset.get("asset_id") or "")
    filename = str(asset.get("filename") or "")
    relative_filepath = str(asset.get("relative_filepath") or "")
    path = resolve_registered_path(project_root, relative_filepath)
    exists = path.is_file()

    result: dict[str, Any] = {
        "asset_id": asset_id,
        "source_image_id": asset.get("source_image_id"),
        "parent_asset_id": asset.get("parent_asset_id"),
        "is_original": asset.get("is_original"),
        "variant_type": asset.get("variant_type"),
        "variant_level": asset.get("variant_level"),
        "filename": filename,
        "relative_filepath": relative_filepath,
        "registry_match_check": registry_id_counts.get(asset_id, 0) == 1,
        "file_exists": exists,
        "file_readable": False,
        "file_extension": path.suffix.lower() if exists else None,
        "detected_format": None,
        "extension_format_match": False,
        "mime_type": mimetypes.guess_type(filename)[0] if exists else None,
        "file_size_bytes": path.stat().st_size if exists else None,
        "width_px": None,
        "height_px": None,
        "aspect_ratio": None,
        "orientation": None,
        "bit_depth": None,
        "color_mode": None,
        "icc_profile": None,
        "exif_available": False,
        "date_captured": None,
        "camera_make": None,
        "camera_model": None,
        "lens_model": None,
        "focal_length_mm": None,
        "iso": None,
        "aperture_f_number": None,
        "shutter_speed_seconds": None,
        "sha256": sha256_file(path) if exists else None,
        "exact_duplicate": False,
        "duplicate_group_id": None,
        "filename_rule_pass": filename_rule_passes(asset),
        "folder_rule_pass": folder_rule_passes(asset),
        "required_metadata_complete_pct": 0.0,
        "intake_status": "fail",
        "failure_reason": None,
        "inspection_timestamp": inspection_timestamp
        or datetime.now().isoformat(sep=" ", timespec="milliseconds"),
        "inspection_method": INSPECTION_METHOD,
        "notes": None,
    }

    if exists:
        try:
            with Image.open(path) as image:
                image.load()
                detected_format = (image.format or "").lower() or None
                width, height = image.size
                exif = image.getexif()

                result.update(
                    {
                        "file_readable": True,
                        "detected_format": detected_format,
                        "mime_type": Image.MIME.get(image.format)
                        or result["mime_type"],
                        "width_px": width,
                        "height_px": height,
                        "aspect_ratio": _aspect_ratio(width, height),
                        "orientation": "landscape"
                        if width > height
                        else ("portrait" if height > width else "square"),
                        "bit_depth": _bit_depth(image.mode),
                        "color_mode": image.mode,
                        "icc_profile": "embedded"
                        if image.info.get("icc_profile")
                        else "none",
                        "exif_available": bool(exif),
                    }
                )
                if exif:
                    result.update(
                        {
                            "date_captured": _exif_scalar(
                                exif.get(36867) or exif.get(36868) or exif.get(306)
                            ),
                            "camera_make": _exif_scalar(exif.get(271)),
                            "camera_model": _exif_scalar(exif.get(272)),
                            "lens_model": _exif_scalar(exif.get(42036)),
                            "focal_length_mm": _exif_scalar(exif.get(37386)),
                            "iso": _exif_scalar(exif.get(34855) or exif.get(34867)),
                            "aperture_f_number": _exif_scalar(exif.get(33437)),
                            "shutter_speed_seconds": _exif_scalar(exif.get(33434)),
                        }
                    )
        except (OSError, SyntaxError, ValueError):
            # The established manifest records failure codes, not exception text.
            pass

    extension = result["file_extension"]
    detected = result["detected_format"]
    accepted_formats = {
        ".jpg": {"jpeg"},
        ".jpeg": {"jpeg"},
        ".png": {"png"},
        ".tif": {"tiff"},
        ".tiff": {"tiff"},
    }
    result["extension_format_match"] = bool(
        extension in accepted_formats and detected in accepted_formats[extension]
    )
    return result


def inspect_assets(
    registry_assets: Sequence[Mapping[str, Any]], project_root: Path
) -> list[dict[str, Any]]:
    """Inspect all registry assets in registry order."""
    counts = Counter(str(asset.get("asset_id") or "") for asset in registry_assets)
    return [inspect_asset(asset, project_root, counts) for asset in registry_assets]


def apply_duplicate_detection(rows: list[dict[str, Any]]) -> None:
    """Assign stable groups and flag every occurrence after the first as duplicate."""
    groups: dict[str, list[int]] = defaultdict(list)
    for index, row in enumerate(rows):
        if row.get("sha256"):
            groups[str(row["sha256"])].append(index)

    duplicate_groups = sorted(
        (indices for indices in groups.values() if len(indices) > 1),
        key=lambda indices: indices[0],
    )
    for group_number, indices in enumerate(duplicate_groups, start=1):
        group_id = f"DUP_{group_number:03d}"
        for position, index in enumerate(indices):
            rows[index]["duplicate_group_id"] = group_id
            rows[index]["exact_duplicate"] = position > 0


def finalize_validation(rows: list[dict[str, Any]]) -> None:
    """Calculate completeness, failure codes, and final intake status."""
    for row in rows:
        populated = sum(
            row.get(field) is not None and row.get(field) != ""
            for field in REQUIRED_METADATA_FIELDS
        )
        row["required_metadata_complete_pct"] = populated / len(
            REQUIRED_METADATA_FIELDS
        )

        failures: list[str] = []
        if not row["file_exists"]:
            failures.append("FILE_MISSING")
        if not row["file_readable"]:
            failures.append("FILE_UNREADABLE")
        if not row["registry_match_check"]:
            failures.append("REGISTRY_MISMATCH")
        if not row["extension_format_match"]:
            failures.append("FORMAT_MISMATCH")
        if not row["filename_rule_pass"]:
            failures.append("INVALID_FILENAME")
        if not row["folder_rule_pass"]:
            failures.append("INVALID_FOLDER")
        if not (
            isinstance(row["width_px"], int)
            and row["width_px"] > 0
            and isinstance(row["height_px"], int)
            and row["height_px"] > 0
        ):
            failures.append("DIMENSIONS_MISSING")
        if not row["sha256"]:
            failures.append("HASH_MISSING")
        if row["exact_duplicate"]:
            failures.append("EXACT_DUPLICATE")
        if row["required_metadata_complete_pct"] < 1:
            failures.append("METADATA_INCOMPLETE")

        row["failure_reason"] = "; ".join(failures) if failures else None
        row["intake_status"] = "fail" if failures else "pass"


def run_intake(project_root: Path, registry_path: Path | None = None) -> list[dict[str, Any]]:
    """Run the complete Stage 1 inspection against all registered assets."""
    registry = registry_path or (
        project_root / "data" / "raw_tables" / "bodyframe_master_registry.xlsx"
    )
    rows = inspect_assets(load_registry_assets(registry), project_root)
    apply_duplicate_detection(rows)
    finalize_validation(rows)
    return rows


def summarize_results(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    """Return the established intake summary measures."""
    total = len(rows)
    completeness = (
        sum(float(row["required_metadata_complete_pct"]) for row in rows) / total
        if total
        else 0.0
    )
    return {
        "total_registered_assets": total,
        "files_found": sum(bool(row["file_exists"]) for row in rows),
        "files_readable": sum(bool(row["file_readable"]) for row in rows),
        "intake_passes": sum(row["intake_status"] == "pass" for row in rows),
        "intake_failures": sum(row["intake_status"] == "fail" for row in rows),
        "exact_duplicates": sum(bool(row["exact_duplicate"]) for row in rows),
        "metadata_completeness": completeness,
        "intake_pass_rate": (
            sum(row["intake_status"] == "pass" for row in rows) / total
            if total
            else 0.0
        ),
        "exif_available": sum(bool(row["exif_available"]) for row in rows),
    }


def _normalise_comparison_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="milliseconds")
    return value


def _values_match(left: Any, right: Any) -> bool:
    left = _normalise_comparison_value(left)
    right = _normalise_comparison_value(right)
    if left is None or right is None:
        return left is right
    if isinstance(left, bool) or isinstance(right, bool):
        return bool(left) is bool(right)
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=0, abs_tol=1e-9)
    return str(left) == str(right)


def compare_with_canonical(
    computed_rows: Sequence[Mapping[str, Any]],
    canonical_rows: Sequence[Mapping[str, Any]],
    ignored_fields: Iterable[str] = ("inspection_timestamp",),
) -> dict[str, Any]:
    """Compare current dry-run values with canonical rows by asset_id."""
    ignored = set(ignored_fields)
    computed = {str(row["asset_id"]): row for row in computed_rows}
    mismatches: list[dict[str, Any]] = []
    matched_rows = 0

    for canonical in canonical_rows:
        asset_id = str(canonical["asset_id"])
        current = computed.get(asset_id)
        if current is None:
            mismatches.append({"asset_id": asset_id, "field": "asset_id", "issue": "missing"})
            continue
        row_matches = True
        for field in MANIFEST_FIELDS:
            if field in ignored:
                continue
            if not _values_match(current.get(field), canonical.get(field)):
                row_matches = False
                mismatches.append(
                    {
                        "asset_id": asset_id,
                        "field": field,
                        "computed": current.get(field),
                        "canonical": canonical.get(field),
                    }
                )
        matched_rows += int(row_matches)

    return {
        "canonical_rows": len(canonical_rows),
        "matched_rows": matched_rows,
        "mismatch_count": len(mismatches),
        "result": "PASS" if not mismatches else "FAIL",
        "mismatches": mismatches,
    }


def write_results_csv(rows: Sequence[Mapping[str, Any]], output_path: Path) -> None:
    """Write a noncanonical result file only when an explicit path is supplied."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def reproducibility_checks(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    """Summarize the four explicitly required Stage 1 reproduction checks."""
    by_id = {str(row["asset_id"]): row for row in rows}
    expansion = [row for row in rows if EXPANSION_ASSET_RE.fullmatch(str(row["asset_id"]))]
    controlled = [by_id[asset_id] for asset_id in CONTROLLED_FAILURE_IDS]
    pilot = [by_id[asset_id] for asset_id in CLEAN_PILOT_IDS]
    original = by_id["BF_S001_0001_ORIGINAL"]
    duplicate = by_id["BF_S001_0001_DUPLICATE"]
    return {
        "expansion": {
            "evaluated": len(expansion),
            "passes": sum(row["intake_status"] == "pass" for row in expansion),
            "failures": sum(row["intake_status"] == "fail" for row in expansion),
        },
        "controlled_failures": {
            "evaluated": len(controlled),
            "detected_as_failures": sum(
                row["intake_status"] == "fail" for row in controlled
            ),
            "failure_reasons": {
                str(row["asset_id"]): row["failure_reason"] for row in controlled
            },
        },
        "clean_pilot": {
            "evaluated": len(pilot),
            "passes": sum(row["intake_status"] == "pass" for row in pilot),
        },
        "known_duplicate": {
            "group_id": duplicate["duplicate_group_id"],
            "original_flagged": bool(original["exact_duplicate"]),
            "controlled_copy_flagged": bool(duplicate["exact_duplicate"]),
            "reproduced": original["duplicate_group_id"] == "DUP_001"
            and duplicate["duplicate_group_id"] == "DUP_001"
            and not original["exact_duplicate"]
            and duplicate["exact_duplicate"],
        },
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="BODYFRAME project root (defaults to the script's parent project).",
    )
    parser.add_argument("--registry", type=Path, help="Optional registry override.")
    parser.add_argument(
        "--canonical-manifest",
        type=Path,
        help="Optional canonical Stage 1 workbook override.",
    )
    parser.add_argument(
        "--compare-canonical",
        action="store_true",
        help="Compare dry-run results with current FileManifestTable rows.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Explicitly confirm that no result file should be written.",
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        help="Optional noncanonical CSV output path. Omit for read-only operation.",
    )
    parser.add_argument("--json", action="store_true", help="Emit JSON summary output.")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.dry_run and args.output_csv:
        parser.error("--dry-run cannot be combined with --output-csv")

    project_root = args.project_root.resolve()
    registry_path = args.registry or (
        project_root / "data" / "raw_tables" / "bodyframe_master_registry.xlsx"
    )
    canonical_path = args.canonical_manifest or (
        project_root / "01_file_intake" / "bodyframe_file_manifest.xlsx"
    )

    rows = run_intake(project_root, registry_path)
    payload: dict[str, Any] = {
        "mode": "dry_run" if not args.output_csv else "write_noncanonical_csv",
        "summary": summarize_results(rows),
        "reproducibility": reproducibility_checks(rows),
    }

    if args.compare_canonical:
        payload["canonical_comparison"] = compare_with_canonical(
            rows, load_canonical_manifest(canonical_path)
        )
    if args.output_csv:
        write_results_csv(rows, args.output_csv)
        payload["output_csv"] = str(args.output_csv)

    if args.json:
        print(json.dumps(payload, indent=2, default=str))
    else:
        print(json.dumps(payload, indent=2, default=str))

    comparison_pass = payload.get("canonical_comparison", {}).get("result", "PASS") == "PASS"
    reproduction = payload["reproducibility"]
    required_pass = (
        reproduction["expansion"] == {"evaluated": 12, "passes": 12, "failures": 0}
        and reproduction["controlled_failures"]["detected_as_failures"] == 5
        and reproduction["clean_pilot"] == {"evaluated": 4, "passes": 4}
        and reproduction["known_duplicate"]["reproduced"]
    )
    return 0 if comparison_pass and required_pass else 1


if __name__ == "__main__":
    raise SystemExit(main())
